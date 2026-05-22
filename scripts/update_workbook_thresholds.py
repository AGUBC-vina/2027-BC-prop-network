"""Add MT/MO/IM_2027/Threshold_Source columns to BC Network 2026 v8.xlsx.

Backs up the original to secondary/BC_Network_2026_v8.backup-before-thresholds.xlsx
the first time it runs (only if no such backup exists yet), then appends four
new columns after the existing data:

    W: MT_ft
    X: MO_ft
    Y: IM_2027_ft
    Z: Threshold_Source     ("2022 GSP" | "AGWL Mirror" | blank for non-RMS)

Values are populated for the 30 thresholded 2027 RMS completions (out of
35 flagged is_2027_gwl_rms); 5 supplemental Chico nested completions and
the 44 non-RMS rows are left blank.
"""
import json
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "BC Network 2026 v8.xlsx"
THRESH_JSON = ROOT / "data" / "thresholds.json"
BACKUP_DIR = ROOT / "secondary"


def main():
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = BACKUP_DIR / "BC_Network_2026_v8.backup-before-thresholds.xlsx"
    if not backup_path.exists():
        shutil.copy2(XLSX, backup_path)
        print(f"Backed up original to {backup_path}")
    else:
        # Restore from backup before writing so the script is idempotent —
        # re-runs always produce the same output rather than appending new
        # threshold column blocks each time.
        shutil.copy2(backup_path, XLSX)
        print(f"Restored {XLSX.name} from backup before writing")

    thresholds = {t["swn"]: t for t in json.loads(THRESH_JSON.read_text())}

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Monitoring Network - 2027 (BC)"]

    # After backup-restore, the workbook has 22 metadata columns; append
    # the 4 threshold columns starting at the next position.
    first_new_col = ws.max_column + 1
    print(f"Writing threshold columns starting at column {first_new_col} ({openpyxl.utils.get_column_letter(first_new_col)})")

    header_font = Font(bold=True)
    headers = ["MT_ft", "MO_ft", "IM_2027_ft", "Threshold_Source"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=1, column=first_new_col + i, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(first_new_col + i)].width = 20 if i == 3 else 12

    mirror_fill = PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")  # warm cream
    adopted_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # light blue

    n_adopted = n_mirror = 0
    for row_idx in range(2, ws.max_row + 1):
        swn = ws.cell(row=row_idx, column=2).value  # column B = "SWN or Well Name"
        if not swn:
            continue
        t = thresholds.get(str(swn).strip())
        if not t:
            continue
        ws.cell(row=row_idx, column=first_new_col + 0, value=t["mt_ft"])
        ws.cell(row=row_idx, column=first_new_col + 1, value=t["mo_ft"])
        ws.cell(row=row_idx, column=first_new_col + 2, value=t["im_2027_ft"])
        ws.cell(row=row_idx, column=first_new_col + 3, value=t["source"])

        fill = adopted_fill if t["source"] == "2022 GSP" else mirror_fill
        for i in range(4):
            ws.cell(row=row_idx, column=first_new_col + i).fill = fill

        if t["source"] == "2022 GSP":
            n_adopted += 1
        else:
            n_mirror += 1

    wb.save(XLSX)
    print(f"Updated {XLSX}")
    print(f"  Wrote {n_adopted} '2022 GSP' rows (adopted carry-overs)")
    print(f"  Wrote {n_mirror} 'AGWL Mirror' rows (computed baselines)")
    print(f"  Total: {n_adopted + n_mirror} threshold rows populated")


if __name__ == "__main__":
    main()
